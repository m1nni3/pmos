import openpyxl
import uuid

CAT_MAP = {
    "Rental Income": "rentalIncome",
    "Levy": "levy",
    "Bond Payment": "bondPayments",
    "Letting Agent": "commission",
    "Municipal": "municipal",
    "Maintenance": "maintenance",
}

PROP_MAP = {
    "Oakdale": "p1000000-0000-0000-0000-000000000001",
    "Malindi": "p1000000-0000-0000-0000-000000000002",
    "Indaba": "p1000000-0000-0000-0000-000000000003",
    "Villeroy": "p1000000-0000-0000-0000-000000000004",
}

YEAR = "2026"
MONTH_COLS = list(range(3, 15))  # C-N in Excel (1-indexed cols 3-14)

wb = openpyxl.load_workbook('/Users/m1nni3/Downloads/2026 Property portfolio (2).xlsx', data_only=True)

STOP_LABELS = {"Budgeted Nett", "Nett Income", "Variance", "Annual"}

def find_section_rows(ws, section_title):
    """Find the data rows under a section title by scanning for known labels."""
    found = {}
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=r, column=2).value
        if cell_val and section_title in str(cell_val):
            for dr in range(1, 30):
                check_row = r + dr
                label = ws.cell(row=check_row, column=2).value
                if not label:
                    continue
                label = str(label).strip()
                if label in STOP_LABELS:
                    break
                if label in CAT_MAP:
                    vals = []
                    for c in MONTH_COLS:
                        v = ws.cell(row=check_row, column=c).value
                        if v is None or v == '' or v == 'N/A' or v == '-':
                            vals.append(0)
                        else:
                            vals.append(int(round(float(v))))
                    found[label] = vals
            break
    return found

def compute_portfolio_total(all_property_data):
    """Sum values across properties per (month, category)."""
    total = {}
    for prop_data in all_property_data:
        for cat_label, monthly_vals in prop_data.items():
            key = cat_label
            if key not in total:
                total[key] = [0] * 12
            for m in range(12):
                total[key][m] += monthly_vals[m]
    return total

all_property_rows = []
for prop_name, prop_id in PROP_MAP.items():
    if prop_name not in wb.sheetnames:
        continue
    ws = wb[prop_name]
    rows = find_section_rows(ws, "Profit/Loss Actual")
    if not rows:
        # try alternate header
        rows = find_section_rows(ws, "Actual")
    all_property_rows.append((prop_id, rows))

print("-- Seed actual data from spreadsheet")
for prop_id, rows in all_property_rows:
    for cat_label, monthly_vals in rows.items():
        cat_key = CAT_MAP[cat_label]
        for m_idx, amt in enumerate(monthly_vals, 1):
            uid = str(uuid.uuid4())
            print(f"INSERT OR REPLACE INTO pl_monthly (id, property_id, year, month, category_key, amount) VALUES ('{uid}', '{prop_id}', '{YEAR}', {m_idx}, '{cat_key}', {amt});")

# Portfolio totals
portfolio = compute_portfolio_total([rows for _, rows in all_property_rows])
for cat_label, monthly_vals in portfolio.items():
    cat_key = CAT_MAP[cat_label]
    for m_idx, amt in enumerate(monthly_vals, 1):
        uid = str(uuid.uuid4())
        print(f"INSERT OR REPLACE INTO pl_monthly (id, property_id, year, month, category_key, amount) VALUES ('{uid}', NULL, '{YEAR}', {m_idx}, '{cat_key}', {amt});")

# Insert budget amounts into pl_data (annual totals from spreadsheet budget section)
print("\n-- Seed budget data into pl_data")
for prop_name, prop_id in PROP_MAP.items():
    if prop_name not in wb.sheetnames:
        continue
    ws = wb[prop_name]
    rows = find_section_rows(ws, "Profit/Loss Budget")
    if not rows:
        rows = find_section_rows(ws, "Budget")
    for cat_label, monthly_vals in rows.items():
        annual = sum(monthly_vals)
        cat_key = CAT_MAP[cat_label]
        uid = str(uuid.uuid4())
        print(f"INSERT OR REPLACE INTO pl_data (id, property_id, year, category, budget_amount, actual_override) VALUES ('{uid}', '{prop_id}', '{YEAR}', '{cat_key}', {annual}, NULL);")

# Portfolio budget totals
portfolio_budget = compute_portfolio_total([rows for _, rows in [(None, find_section_rows(wb[pn], "Profit/Loss Budget") or find_section_rows(wb[pn], "Budget")) for pn in PROP_MAP.keys() if pn in wb.sheetnames]  if rows])
# Simpler: just sum per-property budget totals
print("\n-- Portfolio-level budgets")
total_budget = {}
for prop_name, prop_id in PROP_MAP.items():
    if prop_name not in wb.sheetnames:
        continue
    ws = wb[prop_name]
    rows = find_section_rows(ws, "Profit/Loss Budget")
    if not rows:
        rows = find_section_rows(ws, "Budget")
    for cat_label, monthly_vals in rows.items():
        annual = sum(monthly_vals)
        cat_key = CAT_MAP[cat_label]
        total_budget[cat_key] = total_budget.get(cat_key, 0) + annual

for cat_key, annual in total_budget.items():
    uid = str(uuid.uuid4())
    print(f"INSERT OR REPLACE INTO pl_data (id, property_id, year, category, budget_amount, actual_override) VALUES ('{uid}', NULL, '{YEAR}', '{cat_key}', {annual}, NULL);")
